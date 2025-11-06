import pandas as pd, numpy as np, zipfile, io, re
from pathlib import Path
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

ZIP_PATH = Path("Preparação Análise IA - CSV.zip")  # deixe o .zip na mesma pasta do script
OUT_XLSX = Path("Comparativo_Horizontal_Cabedelo_v2.xlsx")
TEMP_MIN, TEMP_MAX = 15.0, 30.0

def sniff_delim_and_encoding(sample: bytes):
    for enc in ["utf-8", "latin-1", "cp1252"]:
        try:
            text = sample.decode(enc, errors="ignore")
            semi = text.count(";")
            comma = text.count(",")
            return (";" if semi > comma else ","), enc
        except Exception:
            continue
    return ",", "utf-8"

def parse_sensor_id(name: str):
    m = re.search(r"(EL\d+)", name)
    return m.group(1) if m else Path(name).stem

def parse_timestamp_series(raw_series: pd.Series) -> pd.Series:
    # tenta formatos exatos mais comuns primeiro
    fmts = [
        "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M", "%Y/%m/%d %H:%M"
    ]
    for fmt in fmts:
        ts = pd.to_datetime(raw_series, format=fmt, errors="coerce")
        if ts.notna().mean() > 0.9:
            return ts
    # fallback genérico (sem dayfirst)
    return pd.to_datetime(raw_series, errors="coerce", utc=False, infer_datetime_format=True)

assert ZIP_PATH.exists(), f"ZIP não encontrado: {ZIP_PATH}"

frames = []
logs = []

with zipfile.ZipFile(ZIP_PATH, 'r') as zf:
    for info in zf.infolist():
        if info.is_dir() or not info.filename.lower().endswith(".csv"):
            continue
        sensor_id = parse_sensor_id(info.filename)
        data = zf.read(info)
        sep, enc = sniff_delim_and_encoding(data[:4096])
        df = pd.read_csv(io.BytesIO(data), sep=sep, encoding=enc)
        df.columns = [str(c).strip() for c in df.columns]

        # detecta timestamp
        ts_col = None
        for c in df.columns:
            cl = c.lower()
            if any(k in cl for k in ["timestamp","data e hora","data/hora","datahora","data","date","hora","time"]):
                ts_col = c; break
        if ts_col is None: ts_col = df.columns[0]
        ts_raw = df[ts_col].astype(str)
        ts = parse_timestamp_series(ts_raw)

        # detecta temperatura
        t_col = None
        for c in df.columns:
            cl = c.lower()
            if ("temp" in cl) or ("°c" in cl) or (cl.strip() in ["c","tc","temperatura"]):
                t_col = c; break
        if t_col is None:
            candidates = [c for c in df.columns if c != ts_col]
            for c in candidates:
                s = pd.to_numeric(df[c].astype(str).str.replace(",", ".", regex=False), errors="coerce")
                if s.notna().mean() > 0.6:
                    t_col = c; break
            if t_col is None and candidates:
                t_col = candidates[0]
        temp = pd.to_numeric(
            df[t_col].astype(str).str.replace(",", ".", regex=False).str.extract(r"([\-0-9\.]+)")[0],
            errors="coerce"
        )

        cur = pd.DataFrame({
            "timestamp": ts,
            sensor_id: temp,
            "timestamp_original": ts_raw
        })
        # preserva TODAS as medições — apenas descarta onde timestamp não foi parseado
        cur = cur[~cur["timestamp"].isna()].sort_values("timestamp")
        frames.append(cur[["timestamp","timestamp_original",sensor_id]])
        logs.append({"file": info.filename, "sensor_id": sensor_id, "rows_in": int(df.shape[0]), "rows_kept": int(cur.shape[0])})

from functools import reduce
wide_all = reduce(lambda L, R: pd.merge(L, R, on=["timestamp","timestamp_original"], how="outer"), frames)

# se diferentes CSVs tiverem textos levemente distintos para o MESMO instante parseado,
# mantemos 1 linha por timestamp (com um exemplo do original)
example_orig = wide_all.groupby("timestamp", as_index=False)["timestamp_original"].first()
wide = wide_all.drop(columns=["timestamp_original"]).groupby("timestamp", as_index=True).first().sort_index()

# summaries
summ_rows = []
sensor_cols = [c for c in wide.columns if c.startswith("EL")]
for sid in sensor_cols:
    s = wide[sid].dropna()
    if s.empty: 
        continue
    summ_rows.append({
        "sensor_id": sid,
        "n_points": int(s.shape[0]),
        "start": s.index.min(),
        "end": s.index.max(),
        "mean_temp_C": round(float(s.mean()),3),
        "min_temp_C": round(float(s.min()),3),
        "max_temp_C": round(float(s.max()),3),
        "std_temp_C": round(float(s.std(ddof=0)),3),
        "amplitude_C": round(float(s.max()-s.min()),3),
    })
summaries_df = pd.DataFrame(summ_rows).sort_values("sensor_id")

# excursões (≥1 min) — conta blocos contínuos fora da faixa, sem alinhar com outros sensores
TEMP_MIN, TEMP_MAX = 15.0, 30.0
def excursions_from_series(series: pd.Series):
    s = series.dropna()
    if s.empty: return 0, 0, "00:00"
    out = ~s.between(TEMP_MIN, TEMP_MAX)
    idx = s.index
    grp = (out.ne(out.shift()) | (idx.to_series().diff() != pd.Timedelta("1min"))).cumsum()
    count = int(grp[out].nunique())
    total_min = int(out.sum())
    return count, total_min, f"{total_min//60:02d}:{total_min%60:02d}"

exc_rows = []
for sid in sensor_cols:
    cnt, tot_min, hhmm = excursions_from_series(wide[sid])
    n = int(wide[sid].dropna().shape[0])
    pct_out = round(100.0 * (tot_min / max(1, n)), 3)
    exc_rows.append({
        "sensor_id": sid,
        "n_points": n,
        "excursions_count": cnt,
        "excursions_total_min": tot_min,
        "excursions_total_hhmm": hhmm,
        "pct_time_out_of_range_15_30": pct_out,
    })
exc_df = pd.DataFrame(exc_rows).sort_values("sensor_id")

# grava Excel com formatação de data/hora na coluna A
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "comparativo_horizontal"
ws.append(["timestamp"] + sensor_cols)
for ts, row in wide[sensor_cols].iterrows():
    ws.append([ts] + row.tolist())
for cell in ws['A'][1:]:
    cell.number_format = 'yyyy-mm-dd hh:mm:ss'

ws2 = wb.create_sheet("summaries")
for r in dataframe_to_rows(summaries_df, index=False, header=True):
    ws2.append(r)

ws3 = wb.create_sheet("consolidado_excursões")
for r in dataframe_to_rows(exc_df, index=False, header=True):
    ws3.append(r)

ws4 = wb.create_sheet("import_log")
log_df = pd.DataFrame(logs)
for r in dataframe_to_rows(log_df, index=False, header=True):
    ws4.append(r)

ws5 = wb.create_sheet("timestamp_original_exemplo")
ex = example_orig.rename(columns={"timestamp_original":"exemplo_original"})
for r in dataframe_to_rows(ex, index=False, header=True):
    ws5.append(r)

wb.save(OUT_XLSX.as_posix())
print(f"Gerado: {OUT_XLSX.as_posix()}")
