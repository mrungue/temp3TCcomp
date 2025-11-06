"""
Análise de Correlação Térmica - Galpão vs Temperatura Externa
Comparação antes/depois da implementação da tecnologia 3TC

Autor: Análise para cliente
Data: 2024
"""

import pandas as pd
import numpy as np
from pathlib import Path
import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from scipy import stats
from typing import Optional, Dict, List
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURAÇÕES
# ============================================================================
CSV_DIR = Path(".")  # diretório atual com os CSVs
OUTPUT_XLSX = Path("Analise_Correlacao_Termica_3TC.xlsx")
TEMP_MIN, TEMP_MAX = 15.0, 30.0  # faixa ideal
CIDADE = "Cabedelo"  # ajustar conforme necessário
FUSO_HORARIO = "America/Fortaleza"  # UTC-3

# Se você tiver um arquivo CSV com dados meteorológicos externos, defina aqui:
# EXTERNAL_TEMP_CSV = Path("temperatura_externa_cidade.csv")
EXTERNAL_TEMP_CSV = None  # None se precisar buscar via API

# ============================================================================
# FUNÇÕES AUXILIARES
# ============================================================================

def sniff_delim_and_encoding(filepath: Path):
    """Detecta delimitador e encoding do CSV"""
    with open(filepath, 'rb') as f:
        sample = f.read(4096)
        # Tenta encodings brasileiros primeiro (mais comum)
        for enc in ["latin-1", "cp1252", "iso-8859-1", "utf-8"]:
            try:
                text = sample.decode(enc, errors="ignore")
                semi = text.count(";")
                comma = text.count(",")
                return (";" if semi > comma else ","), enc
            except (UnicodeDecodeError, Exception):
                continue
    # Fallback seguro
    return ";", "latin-1"

def parse_sensor_id(filename: str) -> str:
    """Extrai ID do sensor do nome do arquivo"""
    m = re.search(r"(EL\d+)", filename)
    return m.group(1) if m else Path(filename).stem

def parse_timestamp_series(raw_series: pd.Series) -> pd.Series:
    """Parse de timestamps com múltiplos formatos"""
    # Ordem de tentativa: formato ISO primeiro (mais comum nos CSVs dos sensores)
    # depois formatos brasileiros como fallback
    fmts = [
        "%Y-%m-%d %H:%M:%S",  # 2024-02-15 18:00:00 (FORMATO ISO - MAIS COMUM NOS CSVs)
        "%Y-%m-%d %H:%M",     # 2024-02-15 18:00
        "%d/%m/%Y %H:%M:%S",  # 15/02/2024 18:00:00 (formato brasileiro)
        "%d/%m/%Y %H:%M",     # 15/02/2024 18:00 (formato brasileiro)
        "%Y/%m/%d %H:%M:%S",  # 2024/02/15 18:00:00
        "%Y/%m/%d %H:%M",     # 2024/02/15 18:00
    ]
    
    for fmt in fmts:
        try:
            # Para formato ISO não precisa dayfirst, para formato brasileiro sim
            dayfirst = fmt.startswith("%d/")
            ts = pd.to_datetime(raw_series, format=fmt, errors="coerce", utc=False, dayfirst=dayfirst)
            if ts.notna().mean() > 0.9:
                return ts
        except Exception:
            continue
    
    # Fallback: tenta inferir automaticamente (tenta ambos os formatos)
    try:
        # Primeiro tenta sem dayfirst (formato ISO)
        ts = pd.to_datetime(raw_series, errors="coerce", utc=False, dayfirst=False, infer_datetime_format=False)
        if ts.notna().mean() > 0.9:
            return ts
    except Exception:
        pass
    
    try:
        # Depois tenta com dayfirst (formato brasileiro)
        ts = pd.to_datetime(raw_series, errors="coerce", utc=False, dayfirst=True, infer_datetime_format=False)
        if ts.notna().mean() > 0.9:
            return ts
    except Exception:
        pass
    
    return pd.to_datetime(raw_series, errors="coerce", utc=False, dayfirst=False)

def load_external_temperature_data(filepath: Optional[Path] = None) -> Optional[pd.DataFrame]:
    """
    Carrega dados de temperatura externa.
    
    Se filepath fornecido, lê CSV esperando colunas:
    - timestamp (ou data/hora)
    - temperatura (ou temp, temp_externa, etc.)
    
    Se None, retorna None (será necessário buscar via API ou fornecer manualmente)
    """
    if filepath is None or not filepath.exists():
        return None
    
    sep, enc = sniff_delim_and_encoding(filepath)
    
    # Tenta ler com encoding detectado, se falhar tenta outros
    df = None
    encodings_to_try = [enc, "latin-1", "cp1252", "iso-8859-1", "utf-8"]
    for encoding in encodings_to_try:
        try:
            df = pd.read_csv(filepath, sep=sep, encoding=encoding)
            break
        except (UnicodeDecodeError, Exception) as e:
            if encoding == encodings_to_try[-1]:  # Último encoding
                raise e
            continue
    
    if df is None:
        return None
        
    df.columns = [str(c).strip().lower() for c in df.columns]
    
    # Detecta coluna de timestamp
    ts_col = None
    for c in df.columns:
        if any(k in c for k in ["timestamp", "data", "date", "hora", "time"]):
            ts_col = c
            break
    
    # Detecta coluna de temperatura
    temp_col = None
    for c in df.columns:
        if any(k in c for k in ["temp", "temperatura", "celsius", "°c"]):
            if "extern" in c or ts_col is None or c != ts_col:
                temp_col = c
                break
    
    if ts_col is None or temp_col is None:
        print(f"AVISO: Não foi possível detectar colunas de timestamp/temperatura em {filepath}")
        return None
    
    ts = parse_timestamp_series(df[ts_col])
    temp = pd.to_numeric(
        df[temp_col].astype(str).str.replace(",", ".", regex=False).str.extract(r"([\-0-9\.]+)")[0],
        errors="coerce"
    )
    
    result = pd.DataFrame({
        "timestamp": ts,
        "temp_externa": temp
    }).dropna().set_index("timestamp").sort_index()
    
    return result

def get_weather_api_data(start_date: datetime, end_date: datetime, 
                        cidade: str = "Cabedelo") -> Optional[pd.DataFrame]:
    """
    Busca dados meteorológicos via API.
    
    NOTA: Você precisará implementar esta função conforme a API escolhida.
    Opções populares:
    - OpenWeatherMap (requer API key)
    - INMET (Instituto Nacional de Meteorologia - Brasil, gratuito)
    - WeatherAPI
    
    Exemplo de estrutura esperada:
    return pd.DataFrame({
        "timestamp": [...],
        "temp_externa": [...]
    }).set_index("timestamp")
    """
    print(f"AVISO: Função de API não implementada. Use load_external_temperature_data() com arquivo CSV.")
    return None

# ============================================================================
# PROCESSAMENTO DOS DADOS INTERNOS (GALPÃO)
# ============================================================================

def process_internal_sensors(csv_dir: Path) -> tuple:
    """Processa todos os CSVs de sensores internos"""
    frames = []
    logs = []
    
    csv_files = list(csv_dir.glob("*.csv"))
    if not csv_files:
        raise FileNotFoundError(f"Nenhum arquivo CSV encontrado em {csv_dir}")
    
    print(f"Processando {len(csv_files)} arquivos CSV...")
    
    for csv_file in csv_files:
        try:
            sensor_id = parse_sensor_id(csv_file.name)
            sep, enc = sniff_delim_and_encoding(csv_file)
            
            # Tenta ler com encoding detectado, se falhar tenta outros
            df = None
            encodings_to_try = [enc, "latin-1", "cp1252", "iso-8859-1", "utf-8"]
            for encoding in encodings_to_try:
                try:
                    df = pd.read_csv(csv_file, sep=sep, encoding=encoding)
                    break
                except (UnicodeDecodeError, Exception) as e:
                    if encoding == encodings_to_try[-1]:  # Último encoding
                        raise e
                    continue
            
            if df is None or df.empty:
                continue
                
            df.columns = [str(c).strip() for c in df.columns]
            
            # Detecta timestamp (prioriza "Tempo" que é o nome da coluna nos CSVs)
            ts_col = None
            for c in df.columns:
                cl = c.lower().strip()
                # Prioriza "tempo" que é o nome exato da coluna nos CSVs
                if cl == "tempo":
                    ts_col = c
                    break
                elif any(k in cl for k in ["timestamp", "data e hora", "data/hora", "datahora", "data", "date", "hora", "time"]):
                    ts_col = c
                    break
            if ts_col is None:
                ts_col = df.columns[0]
            
            ts_raw = df[ts_col].astype(str)
            ts = parse_timestamp_series(ts_raw)
            
            # Detecta temperatura - PRIORIZA colunas com "Temperatura" ou "temp" e "°C"
            # EVITA colunas de umidade (Umidade%RH)
            t_col = None
            
            # Primeiro: procura especificamente por "Temperatura°C" ou "Temperatura"
            for c in df.columns:
                cl = c.lower().strip()
                # Prioriza colunas que contenham "temperatura" e "°c" ou "c"
                if "temperatura" in cl and ("°c" in cl or "c" in cl):
                    t_col = c
                    break
            
            # Segundo: se não achou, procura por "temp" e "°c"
            if t_col is None:
                for c in df.columns:
                    cl = c.lower().strip()
                    if "temp" in cl and ("°c" in cl or "c" in cl):
                        # Certifica que NÃO é umidade
                        if "umidade" not in cl and "umid" not in cl and "rh" not in cl:
                            t_col = c
                            break
            
            # Terceiro: procura qualquer coluna com "temp" (sem "umidade")
            if t_col is None:
                for c in df.columns:
                    cl = c.lower().strip()
                    if "temp" in cl and "umidade" not in cl and "umid" not in cl and "rh" not in cl:
                        t_col = c
                        break
            
            # Último recurso: procura coluna numérica que NÃO seja umidade
            if t_col is None:
                candidates = [c for c in df.columns if c != ts_col]
                for c in candidates:
                    cl = c.lower().strip()
                    # REJEITA colunas de umidade
                    if "umidade" in cl or "umid" in cl or "rh" in cl:
                        continue
                    s = pd.to_numeric(df[c].astype(str).str.replace(",", ".", regex=False), errors="coerce")
                    if s.notna().mean() > 0.6:
                        # Valida que os valores são razoáveis para temperatura (entre -50 e 100°C)
                        if s.dropna().between(-50, 100).mean() > 0.8:
                            t_col = c
                            break
            
            temp = pd.to_numeric(
                df[t_col].astype(str).str.replace(",", ".", regex=False).str.extract(r"([\-0-9\.]+)")[0],
                errors="coerce"
            )
            
            cur = pd.DataFrame({
                "timestamp": ts,
                sensor_id: temp,
                "timestamp_original": ts_raw
            })
            cur = cur[~cur["timestamp"].isna()].sort_values("timestamp")
            
            frames.append(cur[["timestamp", sensor_id]])
            logs.append({
                "arquivo": csv_file.name,
                "sensor_id": sensor_id,
                "linhas_originais": int(df.shape[0]),
                "linhas_processadas": int(cur.shape[0])
            })
            
        except Exception as e:
            print(f"ERRO ao processar {csv_file.name}: {e}")
            continue
    
    # Merge horizontal
    from functools import reduce
    if not frames:
        raise ValueError("Nenhum dado válido foi processado")
    
    wide = reduce(lambda L, R: pd.merge(L, R, on="timestamp", how="outer"), frames)
    wide = wide.set_index("timestamp").sort_index()
    
    # Calcula média dos sensores (representa temperatura interna geral)
    sensor_cols = [c for c in wide.columns if c.startswith("EL")]
    wide["temp_interna_media"] = wide[sensor_cols].mean(axis=1)
    wide["temp_interna_min"] = wide[sensor_cols].min(axis=1)
    wide["temp_interna_max"] = wide[sensor_cols].max(axis=1)
    wide["temp_interna_std"] = wide[sensor_cols].std(axis=1)
    
    return wide, logs, sensor_cols

# ============================================================================
# ANÁLISES DE CORRELAÇÃO E MÉTRICAS
# ============================================================================

def calculate_correlation_metrics(df: pd.DataFrame, temp_interna_col: str, 
                                  temp_externa_col: str) -> Dict:
    """Calcula métricas de correlação entre temperatura interna e externa"""
    data = df[[temp_interna_col, temp_externa_col]].dropna()
    
    if len(data) < 10:
        return {"erro": "Dados insuficientes para correlação"}
    
    x = data[temp_externa_col].values
    y = data[temp_interna_col].values
    
    # Correlação de Pearson
    pearson_r, pearson_p = stats.pearsonr(x, y)
    
    # Correlação de Spearman (não linear)
    spearman_r, spearman_p = stats.spearmanr(x, y)
    
    # Regressão linear simples
    slope, intercept, r_value, p_value, std_err = stats.linregress(x, y)
    
    # R²
    r_squared = r_value ** 2
    
    # Erro médio
    y_pred = slope * x + intercept
    mae = np.mean(np.abs(y - y_pred))
    rmse = np.sqrt(np.mean((y - y_pred) ** 2))
    
    return {
        "n_amostras": len(data),
        "correlacao_pearson": round(float(pearson_r), 4),
        "p_value_pearson": round(float(pearson_p), 6),
        "correlacao_spearman": round(float(spearman_r), 4),
        "p_value_spearman": round(float(spearman_p), 6),
        "coeficiente_angular": round(float(slope), 4),
        "intercepto": round(float(intercept), 4),
        "r_quadrado": round(float(r_squared), 4),
        "mae_celsius": round(float(mae), 3),
        "rmse_celsius": round(float(rmse), 3),
        "interpretacao": "Forte" if abs(pearson_r) > 0.7 else "Moderada" if abs(pearson_r) > 0.4 else "Fraca"
    }

def calculate_thermal_gradient(df: pd.DataFrame, temp_interna_col: str,
                              temp_externa_col: str) -> pd.DataFrame:
    """Calcula gradiente térmico (diferença entre externa e interna)"""
    data = df[[temp_interna_col, temp_externa_col]].copy().dropna()
    
    data["gradiente_termico"] = data[temp_externa_col] - data[temp_interna_col]
    data["gradiente_percentual"] = (data["gradiente_termico"] / data[temp_externa_col] * 100).round(2)
    
    # Eficiência do isolamento (quanto maior o gradiente, melhor o isolamento)
    # Gradiente positivo = isolamento funcionando (externa mais quente que interna)
    data["eficiencia_isolamento"] = data["gradiente_termico"].apply(
        lambda x: "Excelente" if x > 5 else "Boa" if x > 2 else "Moderada" if x > 0 else "Ineficiente"
    )
    
    return data

def analyze_time_lag(df: pd.DataFrame, temp_interna_col: str,
                     temp_externa_col: str, max_lag_hours: int = 6) -> Dict:
    """Analisa lag temporal entre temperatura externa e interna"""
    data = df[[temp_interna_col, temp_externa_col]].dropna()
    
    if len(data) < 100:
        return {"erro": "Dados insuficientes para análise de lag"}
    
    # Testa diferentes lags
    best_lag = 0
    best_corr = 0
    
    lag_results = []
    for lag_hours in range(-max_lag_hours, max_lag_hours + 1):
        lag_minutes = lag_hours * 60
        temp_externa_shifted = data[temp_externa_col].shift(-lag_minutes)
        corr = temp_externa_shifted.corr(data[temp_interna_col])
        
        if not np.isnan(corr):
            lag_results.append({
                "lag_horas": lag_hours,
                "correlacao": round(float(corr), 4)
            })
            if abs(corr) > abs(best_corr):
                best_corr = corr
                best_lag = lag_hours
    
    return {
        "melhor_lag_horas": best_lag,
        "correlacao_no_lag_otimo": round(float(best_corr), 4),
        "todos_lags": lag_results
    }

def analyze_by_time_period(df: pd.DataFrame, temp_interna_col: str,
                           temp_externa_col: str) -> pd.DataFrame:
    """Analisa correlação por período do dia"""
    data = df[[temp_interna_col, temp_externa_col]].copy().dropna()
    
    # Classifica por período
    data["hora"] = data.index.hour
    data["periodo"] = data["hora"].apply(
        lambda h: "Madrugada (00-06h)" if h < 6 else
                 "Manhã (06-12h)" if h < 12 else
                 "Tarde (12-18h)" if h < 18 else
                 "Noite (18-24h)"
    )
    data["dia_semana"] = data.index.strftime("%A")
    data["dia_noite"] = data["hora"].apply(lambda h: "Dia" if 6 <= h < 18 else "Noite")
    
    # Análise por período
    period_analysis = []
    for periodo in data["periodo"].unique():
        subset = data[data["periodo"] == periodo]
        if len(subset) > 10:
            corr, p = stats.pearsonr(subset[temp_externa_col], subset[temp_interna_col])
            period_analysis.append({
                "periodo": periodo,
                "n_amostras": len(subset),
                "temp_externa_media": round(subset[temp_externa_col].mean(), 2),
                "temp_interna_media": round(subset[temp_interna_col].mean(), 2),
                "gradiente_medio": round((subset[temp_externa_col] - subset[temp_interna_col]).mean(), 2),
                "correlacao": round(float(corr), 4),
                "p_value": round(float(p), 6)
            })
    
    return pd.DataFrame(period_analysis)

def analyze_extreme_conditions(df: pd.DataFrame, temp_interna_col: str,
                              temp_externa_col: str, percentile: float = 0.9) -> Dict:
    """Analisa comportamento em condições extremas (dias mais quentes)"""
    data = df[[temp_interna_col, temp_externa_col]].copy().dropna()
    
    if len(data) < 50:
        return {"erro": "Dados insuficientes"}
    
    # Identifica dias mais quentes (percentil superior)
    threshold = data[temp_externa_col].quantile(percentile)
    extreme_days = data[data[temp_externa_col] >= threshold]
    
    normal_days = data[data[temp_externa_col] < threshold]
    
    # Compara comportamento
    return {
        "percentil_utilizado": percentile,
        "threshold_temp_externa": round(float(threshold), 2),
        "dias_extremos": {
            "n_amostras": len(extreme_days),
            "temp_externa_media": round(float(extreme_days[temp_externa_col].mean()), 2),
            "temp_interna_media": round(float(extreme_days[temp_interna_col].mean()), 2),
            "temp_interna_maxima": round(float(extreme_days[temp_interna_col].max()), 2),
            "gradiente_medio": round(float((extreme_days[temp_externa_col] - extreme_days[temp_interna_col]).mean()), 2),
            "excursoes_acima_30c": int((extreme_days[temp_interna_col] > 30).sum()),
            "pct_excursoes": round(100 * (extreme_days[temp_interna_col] > 30).sum() / len(extreme_days), 2)
        },
        "dias_normais": {
            "n_amostras": len(normal_days),
            "temp_externa_media": round(float(normal_days[temp_externa_col].mean()), 2),
            "temp_interna_media": round(float(normal_days[temp_interna_col].mean()), 2),
            "gradiente_medio": round(float((normal_days[temp_externa_col] - normal_days[temp_interna_col]).mean()), 2)
        }
    }

# ============================================================================
# GERAÇÃO DE RELATÓRIO EXCEL
# ============================================================================

def create_excel_report(wide_internal: pd.DataFrame, external_temp: Optional[pd.DataFrame],
                       sensor_cols: List[str], logs: List[Dict], output_path: Path):
    """Cria relatório Excel completo com todas as análises"""
    
    wb = Workbook()
    
    # Remove planilha padrão
    wb.remove(wb.active)
    
    # ========================================================================
    # Planilha 1: Dados Consolidados
    # ========================================================================
    ws1 = wb.create_sheet("dados_consolidados")
    ws1.append(["timestamp"] + sensor_cols + ["temp_interna_media", "temp_interna_min", "temp_interna_max", "temp_interna_std"])
    
    if external_temp is not None:
        ws1.cell(1, len(sensor_cols) + 5, "temp_externa")
    
    for idx, row in wide_internal.iterrows():
        row_data = [idx] + [row.get(c, None) for c in sensor_cols] + [
            row.get("temp_interna_media", None),
            row.get("temp_interna_min", None),
            row.get("temp_interna_max", None),
            row.get("temp_interna_std", None)
        ]
        if external_temp is not None:
            ext_temp = external_temp.loc[external_temp.index == idx, "temp_externa"]
            row_data.append(ext_temp.iloc[0] if len(ext_temp) > 0 else None)
        ws1.append(row_data)
    
    # Formata coluna de timestamp
    for cell in ws1['A'][1:]:
        cell.number_format = 'yyyy-mm-dd hh:mm:ss'
    
    # ========================================================================
    # Planilha 2: Correlação Principal
    # ========================================================================
    if external_temp is not None:
        # Merge com dados externos
        merged = wide_internal.join(external_temp, how="inner")
        temp_interna_col = "temp_interna_media"
        temp_externa_col = "temp_externa"
        
        corr_metrics = calculate_correlation_metrics(merged, temp_interna_col, temp_externa_col)
        
        ws2 = wb.create_sheet("correlacao_principal")
        ws2.append(["Métrica", "Valor"])
        for key, value in corr_metrics.items():
            ws2.append([key.replace("_", " ").title(), value])
        
        # ========================================================================
        # Planilha 3: Gradiente Térmico
        # ========================================================================
        gradient_data = calculate_thermal_gradient(merged, temp_interna_col, temp_externa_col)
        ws3 = wb.create_sheet("gradiente_termico")
        ws3.append(["timestamp", "temp_externa", "temp_interna", "gradiente_termico", 
                   "gradiente_percentual", "eficiencia_isolamento"])
        for idx, row in gradient_data.iterrows():
            ws3.append([idx, row["temp_externa"], row["temp_interna_media"], 
                       row["gradiente_termico"], row["gradiente_percentual"], 
                       row["eficiencia_isolamento"]])
        for cell in ws3['A'][1:]:
            cell.number_format = 'yyyy-mm-dd hh:mm:ss'
        
        # ========================================================================
        # Planilha 4: Análise de Lag Temporal
        # ========================================================================
        lag_analysis = analyze_time_lag(merged, temp_interna_col, temp_externa_col)
        ws4 = wb.create_sheet("analise_lag_temporal")
        if "erro" not in lag_analysis:
            ws4.append(["Métrica", "Valor"])
            for key, value in lag_analysis.items():
                if key != "todos_lags":
                    ws4.append([key.replace("_", " ").title(), value])
            ws4.append([])
            ws4.append(["Lag (horas)", "Correlação"])
            for lag_item in lag_analysis.get("todos_lags", []):
                ws4.append([lag_item["lag_horas"], lag_item["correlacao"]])
        else:
            ws4.append(["Erro", lag_analysis["erro"]])
        
        # ========================================================================
        # Planilha 5: Análise por Período
        # ========================================================================
        period_analysis = analyze_by_time_period(merged, temp_interna_col, temp_externa_col)
        ws5 = wb.create_sheet("analise_por_periodo")
        if not period_analysis.empty:
            for r in dataframe_to_rows(period_analysis, index=False, header=True):
                ws5.append(r)
        
        # ========================================================================
        # Planilha 6: Condições Extremas
        # ========================================================================
        extreme_analysis = analyze_extreme_conditions(merged, temp_interna_col, temp_externa_col)
        ws6 = wb.create_sheet("condicoes_extremas")
        if "erro" not in extreme_analysis:
            ws6.append(["Análise", "Valor"])
            ws6.append(["Percentil Utilizado", extreme_analysis["percentil_utilizado"]])
            ws6.append(["Threshold Temperatura Externa (°C)", extreme_analysis["threshold_temp_externa"]])
            ws6.append([])
            ws6.append(["DIAS EXTREMOS (Mais Quentes)"])
            for key, value in extreme_analysis["dias_extremos"].items():
                ws6.append([key.replace("_", " ").title(), value])
            ws6.append([])
            ws6.append(["DIAS NORMAIS"])
            for key, value in extreme_analysis["dias_normais"].items():
                ws6.append([key.replace("_", " ").title(), value])
        else:
            ws6.append(["Erro", extreme_analysis["erro"]])
    
    # ========================================================================
    # Planilha 7: Log de Importação
    # ========================================================================
    ws7 = wb.create_sheet("log_importacao")
    log_df = pd.DataFrame(logs)
    for r in dataframe_to_rows(log_df, index=False, header=True):
        ws7.append(r)
    
    # ========================================================================
    # Planilha 8: Resumo Estatístico por Sensor
    # ========================================================================
    ws8 = wb.create_sheet("resumo_estatistico_sensores")
    summ_rows = []
    for sid in sensor_cols:
        s = wide_internal[sid].dropna()
        if s.empty:
            continue
        summ_rows.append({
            "sensor_id": sid,
            "n_pontos": int(s.shape[0]),
            "inicio": s.index.min(),
            "fim": s.index.max(),
            "media_c": round(float(s.mean()), 3),
            "min_c": round(float(s.min()), 3),
            "max_c": round(float(s.max()), 3),
            "std_c": round(float(s.std(ddof=0)), 3),
            "amplitude_c": round(float(s.max() - s.min()), 3),
        })
    summaries_df = pd.DataFrame(summ_rows).sort_values("sensor_id")
    for r in dataframe_to_rows(summaries_df, index=False, header=True):
        ws8.append(r)
    
    # Salva arquivo
    wb.save(output_path)
    print(f"\n✓ Relatório Excel gerado: {output_path}")

# ============================================================================
# MAIN
# ============================================================================

def main():
    print("=" * 70)
    print("ANÁLISE DE CORRELAÇÃO TÉRMICA - GALPÃO vs TEMPERATURA EXTERNA")
    print("Comparação antes/depois da tecnologia 3TC")
    print("=" * 70)
    
    # 1. Processa dados internos (galpão)
    print("\n[1/4] Processando dados dos sensores internos...")
    wide_internal, logs, sensor_cols = process_internal_sensors(CSV_DIR)
    print(f"✓ {len(sensor_cols)} sensores processados")
    print(f"✓ {len(wide_internal)} medições temporais")
    
    # 2. Carrega dados externos
    print("\n[2/4] Carregando dados de temperatura externa...")
    external_temp = None
    
    if EXTERNAL_TEMP_CSV and EXTERNAL_TEMP_CSV.exists():
        external_temp = load_external_temperature_data(EXTERNAL_TEMP_CSV)
        if external_temp is not None:
            print(f"✓ Dados externos carregados: {len(external_temp)} medições")
        else:
            print("⚠ Não foi possível carregar dados externos do CSV")
    else:
        print("⚠ Arquivo CSV de temperatura externa não encontrado")
        print("  Para análise completa, forneça um arquivo CSV com:")
        print("    - Coluna de timestamp/data")
        print("    - Coluna de temperatura externa")
        print("  Ou implemente a função get_weather_api_data() para buscar via API")
    
    # 3. Se houver dados externos, faz merge
    if external_temp is not None:
        merged = wide_internal.join(external_temp, how="inner")
        print(f"\n[3/4] Dados alinhados: {len(merged)} medições comuns")
        
        # Análises rápidas
        if len(merged) > 10:
            corr_metrics = calculate_correlation_metrics(merged, "temp_interna_media", "temp_externa")
            print(f"\n📊 CORRELAÇÃO TEMPERATURA EXTERNA vs INTERNA:")
            print(f"   Correlação de Pearson: {corr_metrics.get('correlacao_pearson', 'N/A')}")
            print(f"   R²: {corr_metrics.get('r_quadrado', 'N/A')}")
            print(f"   Interpretação: {corr_metrics.get('interpretacao', 'N/A')}")
            
            gradient = calculate_thermal_gradient(merged, "temp_interna_media", "temp_externa")
            grad_medio = gradient["gradiente_termico"].mean()
            print(f"\n🌡️  GRADIENTE TÉRMICO MÉDIO:")
            print(f"   Temperatura Externa - Interna: {grad_medio:.2f}°C")
            print(f"   (Valores positivos indicam isolamento funcionando)")
    else:
        print("\n⚠ Análises de correlação serão limitadas sem dados de temperatura externa")
    
    # 4. Gera relatório Excel
    print("\n[4/4] Gerando relatório Excel...")
    create_excel_report(wide_internal, external_temp, sensor_cols, logs, OUTPUT_XLSX)
    
    print("\n" + "=" * 70)
    print("ANÁLISE CONCLUÍDA!")
    print("=" * 70)
    print(f"\n📁 Arquivo gerado: {OUTPUT_XLSX}")
    print("\n💡 PRÓXIMOS PASSOS:")
    print("   1. Obtenha dados de temperatura externa da cidade para o período")
    print("   2. Salve em CSV com colunas: timestamp, temperatura")
    print("   3. Configure EXTERNAL_TEMP_CSV no script e execute novamente")
    print("   4. Após implementar tecnologia 3TC, repita a análise para comparação")
    print()

if __name__ == "__main__":
    main()

